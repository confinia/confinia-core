#!/usr/bin/env python3
"""French historical population per commune (issue #88).

Source: INSEE "Historique des populations communales", one Excel workbook with
37 census years (1876-2023). Licence Ouverte 2.0.
  https://www.insee.fr/fr/statistiques/3698339

IMPORTANT — the figures are HARMONISED, not historical. INSEE back-projects
every census onto the geography in effect on a single reference date (1 January
2025 for the 2023 edition): a commune that merged in 2019 carries the SUM of its
constituents back to 1876, and a commune that has disappeared is simply absent
from the file. We therefore store `harmonised_on` and the API/report must SAY SO.
Serving these figures as if they were the historical population of a historical
commune would break the one thing we sell: per-fact provenance.

Coverage: mainland 1876-2023, Corsica from 1936, DOM from 1954/1962, Mayotte
excluded. Empty cells (before a territory entered the series) are skipped.

Note for readers of the data: INSEE warns that censuses from 2006 onward should
only be compared across gaps of at least 5 years (rolling-survey methodology).

Usage:
  ingest_pop.py --xlsx /data/raw/insee-pop/base-pop-historiques-1876-2023.xlsx
                [--geography 2025-01-01] [--dsn $PG_DSN]
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from datetime import date

SOURCE = "insee-pop"
YEAR_LABEL = re.compile(r"Population en (\d{4})")
# The workbook's single data sheet; its name carries the edition span.
SHEET_PREFIX = "pop_"


def read_workbook(path: str) -> tuple[list[int], list[tuple[str, str, list]]]:
    """Return (census_years, rows) where each row is (code, name, values)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    sheets = [s for s in wb.sheetnames if s.startswith(SHEET_PREFIX)]
    if not sheets:
        raise SystemExit(f"No '{SHEET_PREFIX}*' sheet in {path}: unexpected layout.")
    ws = wb[sheets[0]]
    rows = ws.iter_rows(values_only=True)

    # The header is the first row whose cells carry "Population en <year>"; the
    # next row repeats them as variable codes (PMUN…/PSDC…) and is skipped.
    years: list[int] = []
    year_cols: list[int] = []
    code_col = name_col = None
    for row in rows:
        cells = [str(c) if c is not None else "" for c in row]
        hits = [(i, YEAR_LABEL.match(c)) for i, c in enumerate(cells)]
        hits = [(i, m) for i, m in hits if m]
        if hits:
            year_cols = [i for i, _ in hits]
            years = [int(m.group(1)) for _, m in hits]
            code_col = next(i for i, c in enumerate(cells) if "Code" in c)
            name_col = next(i for i, c in enumerate(cells) if "Libellé" in c)
            next(rows)                      # variable-code row (PMUN2023, …)
            break
    if not years:
        raise SystemExit(f"No 'Population en <year>' header found in {path}.")

    out: list[tuple[str, str, list]] = []
    for row in rows:
        code = str(row[code_col]).strip() if row[code_col] is not None else ""
        if not re.fullmatch(r"[0-9AB]{5}", code):
            continue                        # totals / blank tail lines
        name = str(row[name_col] or "").strip()
        if "�" in name:
            raise SystemExit(f"Corrupted source (U+FFFD) at {code}: refusing to ingest.")
        out.append((code, name, [row[i] for i in year_cols]))
    return years, out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="INSEE historical population workbook")
    ap.add_argument("--geography", default="2025-01-01",
                    help="reference date the series are harmonised on (INSEE edition)")
    ap.add_argument("--dsn", default=os.environ.get("PG_DSN"))
    args = ap.parse_args()
    if not args.dsn:
        raise SystemExit("PG_DSN missing (or --dsn).")
    harmonised_on = date.fromisoformat(args.geography)

    years, rows = read_workbook(args.xlsx)
    print(f"{len(rows)} communes, {len(years)} census years "
          f"({min(years)}-{max(years)}), harmonised on {harmonised_on}")

    import psycopg2
    conn = psycopg2.connect(args.dsn)
    inserted = 0
    with conn, conn.cursor() as cur:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS commune_population (
                country       text NOT NULL DEFAULT 'FR',
                code          text NOT NULL,
                census_year   int  NOT NULL,
                population    int  NOT NULL,
                source        text NOT NULL,
                harmonised_on date,
                PRIMARY KEY (country, code, census_year)
            );
            CREATE INDEX IF NOT EXISTS idx_pop_code ON commune_population (country, code);
        """)
        batch = []
        for code, _name, values in rows:
            for year, value in zip(years, values):
                if value is None or str(value).strip() in ("", "N/A", "-"):
                    continue                # territory not in the series that year
                try:
                    pop = int(float(str(value).replace(" ", "").replace(" ", "")))
                except ValueError:
                    continue
                batch.append(("FR", code, year, pop, SOURCE, harmonised_on))
        # Idempotent: replaying the same edition updates in place.
        cur.executemany(
            "INSERT INTO commune_population "
            " (country, code, census_year, population, source, harmonised_on) "
            "VALUES (%s,%s,%s,%s,%s,%s) "
            "ON CONFLICT (country, code, census_year) DO UPDATE SET "
            " population = EXCLUDED.population, source = EXCLUDED.source, "
            " harmonised_on = EXCLUDED.harmonised_on", batch)
        inserted = len(batch)
        cur.execute("SELECT count(*), count(DISTINCT code) FROM commune_population "
                    "WHERE country='FR' AND source=%s", (SOURCE,))
        total, communes = cur.fetchone()
    conn.close()

    print(f"{inserted} figures written; {total} live for {communes} communes")
    # Control: every current commune should carry at least the latest census.
    ok = communes > 30000 and total > 900000
    print(f"control: {'OK' if ok else 'MISMATCH (expected >30k communes)'}")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
